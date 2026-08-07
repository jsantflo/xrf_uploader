"""
compare_calsets.py — Side-by-side comparison of N XRF calibration sets.

Each argument is either a single .txt file or a bracket-quoted group of files
treated as repeated measurements of the same calset.  Groups are averaged;
standard deviation is shown as error bars in Overview Graphs and as CV%
colour-coding in the Overview table.

Usage (quote brackets so the shell doesn't expand them):
    python compare_calsets.py file1.txt file2.txt
    python compare_calsets.py "[file1.txt,file2.txt]" file3 .txt
    python compare_calsets.py "[a.txt,b.txt]" "[c.txt,d.txt]" --output out.xlsx
"""

import argparse
import math
import pathlib
import re
import shutil
import sys

import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from xrf_upload import (
    parse_xrf_txt,
    TEMPLATE_PATH,
    _CALDAT_COLS,
    ELEMENT_CONFIG,
    ELEMENT_FULL_NAMES,
    EXTRA_ELEMENTS,
    YP50_TRUE_CONC,
)


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# First spare column in element sheets used to store per-group SD values
# (cols 50, 51, 52, … — well clear of the stats table at cols 11-18)
_SD_COL_BASE = 50

# Maps YP50-N standard key → element-sheet row number (row 2 = YP50-0, …)
_STD_TO_ROW = {f"YP50-{i}": i + 2 for i in range(7)}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


def _signal_col(si: int) -> int:
    """1-indexed signal column for group index si.
    si=0→8(H), si=1→9(I), si=2→10(J), si=3→19(S), …
    Skips cols 11-18 (template stats table).
    """
    return 8 + si if si <= 2 else 16 + si


def _ref_rows(formula: str):
    m = re.search(r'\$[A-Z]+\$(\d+):\$[A-Z]+\$(\d+)', formula)
    return (int(m.group(1)), int(m.group(2))) if m else (2, 8)


def _col_letter(n: int) -> str:
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _autosize_columns(ws, min_width: int = 10, max_width: int = 40):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _unique_label(base: str, stem: str, seen: set) -> str:
    base = (base or stem)[:28]
    label, n = base, 1
    while label in seen:
        n += 1
        label = f"{base} ({n})"
    seen.add(label)
    return label


# ---------------------------------------------------------------------------
# Statistical helpers
# ---------------------------------------------------------------------------

def _betacf(a, b, x):
    """Continued-fraction worker for _betainc (Numerical Recipes §6.4)."""
    MAXIT, EPS, FPMIN = 200, 3e-7, 1e-300
    qab, qap, qam = a + b, a + 1, a - 1
    c, d = 1.0, 1.0 - qab * x / qap
    if abs(d) < FPMIN: d = FPMIN
    d = 1.0 / d
    h = d
    for m in range(1, MAXIT + 1):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d;  c = 1.0 + aa / c
        if abs(d) < FPMIN: d = FPMIN
        if abs(c) < FPMIN: c = FPMIN
        d = 1.0 / d;  h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d;  c = 1.0 + aa / c
        if abs(d) < FPMIN: d = FPMIN
        if abs(c) < FPMIN: c = FPMIN
        d = 1.0 / d;  delta = d * c;  h *= delta
        if abs(delta - 1.0) < EPS: break
    return h


def _betainc(a, b, x):
    """Regularized incomplete beta function I(x; a, b) — pure stdlib."""
    if x <= 0: return 0.0
    if x >= 1: return 1.0
    lbeta = math.lgamma(a) + math.lgamma(b) - math.lgamma(a + b)
    if x < (a + 1) / (a + b + 2):
        return _betacf(a, b, x) * math.exp(a * math.log(x) + b * math.log(1 - x) - lbeta) / a
    else:
        return 1.0 - _betacf(b, a, 1 - x) * math.exp(b * math.log(1 - x) + a * math.log(x) - lbeta) / b


def _t_pval(t_stat, df):
    """Two-tailed p-value from t-distribution."""
    if df <= 0:
        return None
    return _betainc(df / 2, 0.5, df / (df + t_stat ** 2))


def _f_pval(f_stat, df1, df2):
    """Two-tailed p-value from F-distribution."""
    p_tail = _betainc(df2 / 2, df1 / 2, df2 / (df2 + df1 * f_stat))
    return 2 * min(p_tail, 1 - p_tail)


def _welch_ttest(ys_a, ys_b):
    """Independent-samples Welch t-test (unequal variance). Returns (t, p) or (None, None)."""
    ys_a = [y for y in ys_a if y is not None]
    ys_b = [y for y in ys_b if y is not None]
    na, nb = len(ys_a), len(ys_b)
    if na < 2 or nb < 2:
        return None, None
    ma  = sum(ys_a) / na
    mb  = sum(ys_b) / nb
    sa2 = sum((y - ma) ** 2 for y in ys_a) / (na - 1)
    sb2 = sum((y - mb) ** 2 for y in ys_b) / (nb - 1)
    se  = math.sqrt(sa2 / na + sb2 / nb)
    if se < 1e-12:
        return None, None
    t  = (ma - mb) / se
    df = (sa2 / na + sb2 / nb) ** 2 / ((sa2 / na) ** 2 / (na - 1) + (sb2 / nb) ** 2 / (nb - 1))
    return t, _t_pval(t, df)


def _linreg_full(xs, ys):
    """OLS regression; returns (m, b, SE_m, SE_b, MSE, Sxx, n) or None."""
    n = len(xs)
    if n < 3:
        return None
    xm = sum(xs) / n
    ym = sum(ys) / n
    Sxx = sum((x - xm) ** 2 for x in xs)
    if abs(Sxx) < 1e-12:
        return None
    Sxy = sum((x - xm) * (y - ym) for x, y in zip(xs, ys))
    m = Sxy / Sxx
    b = ym - m * xm
    SSE = sum((y - m * x - b) ** 2 for x, y in zip(xs, ys))
    MSE = SSE / (n - 2)
    SE_m = math.sqrt(MSE / Sxx)
    SE_b = math.sqrt(MSE * (1 / n + xm ** 2 / Sxx))
    return m, b, SE_m, SE_b, MSE, Sxx, n


def _compare_slopes(xs, ys_a, ys_b):
    """t-test H0: slopes of two OLS regressions on the same x grid are equal."""
    ra = _linreg_full(xs, ys_a)
    rb = _linreg_full(xs, ys_b)
    if ra is None or rb is None:
        return None, None
    ma, _ba, SE_ma, _SEba, _MSEa, _Sxxa, na = ra
    mb, _bb, SE_mb, _SEbb, _MSEb, _Sxxb, nb = rb
    se = math.sqrt(SE_ma ** 2 + SE_mb ** 2)
    if se < 1e-12:
        return float("inf"), 0.0
    t = (ma - mb) / se
    return t, _t_pval(t, na + nb - 4)


def _compare_intercepts(xs, ys_a, ys_b):
    """t-test H0: intercepts of two OLS regressions are equal."""
    ra = _linreg_full(xs, ys_a)
    rb = _linreg_full(xs, ys_b)
    if ra is None or rb is None:
        return None, None
    _ma, ba, _SEma, SE_ba, _MSEa, _Sxxa, na = ra
    _mb, bb, _SEmb, SE_bb, _MSEb, _Sxxb, nb = rb
    se = math.sqrt(SE_ba ** 2 + SE_bb ** 2)
    if se < 1e-12:
        return float("inf"), 0.0
    t = (ba - bb) / se
    return t, _t_pval(t, na + nb - 4)


def _precision_ftest(sd_a, n_a, sd_b, n_b, elem, standards):
    """Pooled F-test comparing within-run MSE between two groups (both need n>=2)."""
    if n_a < 2 or n_b < 2:
        return None, None
    def _pool(sd_dict, n):
        vals = [sd_dict.get(s, {}).get(elem) for s in standards]
        vals = [v for v in vals if v is not None]
        ss  = sum(v ** 2 * (n - 1) for v in vals)
        df  = (n - 1) * len(vals)
        return ss, df
    ss_a, df_a = _pool(sd_a, n_a)
    ss_b, df_b = _pool(sd_b, n_b)
    if df_a == 0 or df_b == 0:
        return None, None
    MSE_a, MSE_b = ss_a / df_a, ss_b / df_b
    if MSE_a == 0 and MSE_b == 0:
        return 1.0, 1.0
    if MSE_b == 0 or MSE_a == 0:
        return float("inf"), 0.0
    if MSE_a >= MSE_b:
        f, df1, df2 = MSE_a / MSE_b, df_a, df_b
    else:
        f, df1, df2 = MSE_b / MSE_a, df_b, df_a
    return f, _f_pval(f, df1, df2)


def _paired_ttest(ys_a, ys_b):
    """Paired t-test on per-standard signal differences (fallback when n=1)."""
    diffs = [a - b for a, b in zip(ys_a, ys_b) if a is not None and b is not None]
    n = len(diffs)
    if n < 2:
        return None, None
    md = sum(diffs) / n
    sd = math.sqrt(sum((d - md) ** 2 for d in diffs) / (n - 1))
    if sd < 1e-12:
        return float("inf"), 0.0
    t = md / (sd / math.sqrt(n))
    return t, _t_pval(t, n - 1)


# ---------------------------------------------------------------------------
# Group parsing
# ---------------------------------------------------------------------------

def parse_file_groups(raw_args: list) -> list:
    """
    Convert CLI tokens into a list of Path-lists (groups).

    Recognised syntaxes (brackets MUST be shell-quoted):
      file.txt                      → single-file group
      "[a.txt,b.txt]"               → two-file group (comma-separated)
      "[a.txt b.txt]"               → two-file group (space-separated)
      "[a.txt" "b.txt]"             → same group split across tokens
    """
    groups: list = []
    current: list = None

    for token in raw_args:
        token = token.strip()
        started = token.startswith("[")
        ended   = token.endswith("]")

        if started and ended:
            inner = token[1:-1]
            files = [f.strip() for f in inner.split(",") if f.strip()]
            if files:
                groups.append([pathlib.Path(f) for f in files])
        elif started:
            inner = token[1:]
            current = [pathlib.Path(inner)] if inner else []
        elif ended:
            inner = token[:-1]
            if current is not None:
                if inner:
                    current.append(pathlib.Path(inner))
                groups.append(current)
                current = None
        elif current is not None:
            current.append(pathlib.Path(token))
        else:
            groups.append([pathlib.Path(token)])

    if current:
        groups.extend([[p] for p in current])

    return groups


# ---------------------------------------------------------------------------
# Per-group statistics
# ---------------------------------------------------------------------------

def _compute_group_stats(paths: list):
    """
    Parse one or more .txt files and return averaged calset statistics.

    Returns (mean_cal, sd_cal, label_base, n) where:
      mean_cal / sd_cal : {std_key: {elem: float}}
      label_base        : suggested sheet label string
      n                 : number of valid files in the group
    """
    all_cal_sets = []
    run_dates    = []

    for p in paths:
        parsed   = parse_xrf_txt(p)
        cal_sets = parsed.get("cal_sets", [])
        all_cal_sets.extend(cal_sets)
        rd = parsed.get("run_date") or ""
        if rd:
            run_dates.append(rd)

    if not all_cal_sets:
        return None, None, "", 0

    n = len(all_cal_sets)

    # Build superset of standards and elements
    all_stds  = [k for k in all_cal_sets[0] if not k.startswith("_")]
    all_elems: set = set()
    for cs in all_cal_sets:
        for std in all_stds:
            all_elems.update(cs.get(std, {}).keys())

    mean_cal: dict = {}
    sd_cal:   dict = {}

    for std in all_stds:
        mean_cal[std] = {}
        sd_cal[std]   = {}
        for elem in all_elems:
            vals = [cs.get(std, {}).get(elem) for cs in all_cal_sets]
            vals = [v for v in vals if v is not None]
            if not vals:
                continue
            m = sum(vals) / len(vals)
            mean_cal[std][elem] = m
            if len(vals) > 1:
                var = sum((v - m) ** 2 for v in vals) / (len(vals) - 1)
                sd_cal[std][elem] = math.sqrt(var)
            else:
                sd_cal[std][elem] = 0.0

    # Build label base
    unique_dates = list(dict.fromkeys(run_dates))
    if n == 1:
        label_base = unique_dates[0] if unique_dates else paths[0].stem
    elif len(unique_dates) == 1:
        label_base = f"{unique_dates[0]} (n={n})"
    elif unique_dates:
        label_base = f"{unique_dates[0]}… (n={n})"
    else:
        label_base = f"{paths[0].stem} (n={n})"

    return mean_cal, sd_cal, label_base, n


# ---------------------------------------------------------------------------
# Core builder
# ---------------------------------------------------------------------------

def build_comparison(groups: list, output: pathlib.Path):
    """
    groups : list of Path-lists (each inner list is one calset / group).
    output : destination .xlsx path.
    """
    # --- Compute stats for every group ---
    group_data  = []
    group_paths = []   # parallel list of Path-lists
    seen_labels: set = set()

    for paths in groups:
        mean_cal, sd_cal, label_base, n = _compute_group_stats(paths)
        if mean_cal is None:
            names = [p.name for p in paths]
            print(f"  Warning: no calsets found in {names} — skipping.")
            continue
        label = _unique_label(label_base, paths[0].stem, seen_labels)
        files_desc = ", ".join(p.name for p in paths)
        print(f"  Group '{label}' (n={n}) ← {files_desc}")
        group_data.append((label, mean_cal, sd_cal, n))
        group_paths.append(paths)

    if not group_data:
        sys.exit("Error: no valid calibration data found.")

    n_groups   = len(group_data)
    set_labels = [g[0] for g in group_data]

    # --- Start from template ---
    shutil.copy2(TEMPLATE_PATH, output)
    wb = openpyxl.load_workbook(output)
    del wb["Samdat"]

    # --- Caldat sheets: one per group, filled with mean signals ---
    ws_cal_orig   = wb["Caldat"]
    caldat_sheets = [ws_cal_orig]
    for si in range(1, n_groups):
        ws_copy = wb.copy_worksheet(ws_cal_orig)
        ws_copy.title = set_labels[si]
        caldat_sheets.append(ws_copy)
    ws_cal_orig.title = set_labels[0]

    for ws_c, (label, mean_cal, sd_cal, n) in zip(caldat_sheets, group_data):
        for std_idx, elem_col, sig_col in _CALDAT_COLS:
            std_key = f"YP50-{std_idx}"
            signals = mean_cal.get(std_key, {})
            for row in range(2, ws_c.max_row + 1):
                elem_val = ws_c.cell(row, elem_col).value
                if elem_val and isinstance(elem_val, str):
                    elem_val = elem_val.strip()
                    if elem_val in signals:
                        ws_c.cell(row, sig_col, signals[elem_val])
        _autosize_columns(ws_c)

    # --- Element sheets: signal columns, SD storage, formula re-point ---
    for elem in ELEMENT_CONFIG:
        ws = wb[elem]

        ws.cell(1, 8, f'=""&$B$2&" Signal"&" - {set_labels[0]}"')

        if n_groups >= 2:
            chart_ranges = [
                _ref_rows(ch.series[0].xVal.numRef.f)
                for ch in ws._charts[:3]
            ]
            for si in range(1, n_groups):
                col_sig = _signal_col(si)
                ws.cell(1, col_sig, f'=""&$B$2&" Signal"&" - {set_labels[si]}"')
                for row in range(2, 9):
                    h_val = ws.cell(row, 8).value
                    if isinstance(h_val, str) and h_val.startswith("="):
                        ws.cell(row, col_sig).value = h_val.replace(
                            "Caldat!", f"'{set_labels[si]}'!"
                        )
                    ws.cell(row, col_sig).fill = _YELLOW
                for chart_i, (rmin, rmax) in enumerate(chart_ranges):
                    chart = ws._charts[chart_i]
                    x_ref = Reference(ws, min_col=7,       min_row=rmin, max_row=rmax)
                    y_ref = Reference(ws, min_col=col_sig, min_row=rmin, max_row=rmax)
                    chart.series.append(Series(y_ref, xvalues=x_ref, title=set_labels[si]))

        # Re-point Caldat! → first group's sheet
        label0 = f"'{set_labels[0]}'!"
        for ws_row in ws.iter_rows():
            for cell in ws_row:
                if isinstance(cell.value, str) and "Caldat!" in cell.value:
                    cell.value = cell.value.replace("Caldat!", label0)

        # Write SD values into spare columns (used as error-bar data sources)
        for si, (label, mean_cal, sd_cal, n) in enumerate(group_data):
            if n < 2:
                continue
            sd_col = _SD_COL_BASE + si
            for std_key, row_num in _STD_TO_ROW.items():
                sd_val = sd_cal.get(std_key, {}).get(elem)
                if sd_val is not None:
                    ws.cell(row_num, sd_col, sd_val)

        _autosize_columns(ws)

    # --- Overview Graphs: scatter charts with optional SD error bars ---
    ws_ov = wb.create_sheet("Overview Graphs", 0)
    ws_ov.merge_cells("A1:N1")
    ws_ov.cell(1, 1).value = "XRF Calibration Comparison — " + ", ".join(set_labels)
    ws_ov.cell(1, 1).font  = Font(bold=True, size=13)

    CPR        = 3
    CHART_COLS = 14
    CHART_ROWS = 22

    for chart_i, elem in enumerate(ELEMENT_CONFIG):
        ws_elem = wb[elem]
        ri, ci  = divmod(chart_i, CPR)
        anchor  = f"{_col_letter(1 + ci * CHART_COLS)}{3 + ri * CHART_ROWS}"

        chart = ScatterChart()
        chart.title             = ELEMENT_FULL_NAMES[elem]
        chart.scatterStyle      = "marker"
        chart.x_axis.title      = "True Concentration (ppm)"
        chart.x_axis.numFmt     = "0"
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.delete     = False
        chart.y_axis.title      = "Signal"
        chart.y_axis.numFmt     = "0"
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.delete     = False
        chart.width  = 12
        chart.height = 9

        x_ref = Reference(ws_elem, min_col=7, min_row=2, max_row=8)

        for si, (label, mean_cal, sd_cal, n) in enumerate(group_data):
            col_sig = _signal_col(si)
            y_ref   = Reference(ws_elem, min_col=col_sig, min_row=2, max_row=8)
            s = Series(y_ref, xvalues=x_ref, title=label)
            s.graphicalProperties.line.width = 9525  # 0.75 pt
            s.marker.symbol = "circle"
            s.marker.size   = 8

            if n >= 2:
                sd_col  = _SD_COL_BASE + si
                col_l   = get_column_letter(sd_col)
                err_rng = f"'{elem}'!${col_l}$2:${col_l}$8"
                s.errBars = ErrorBars(
                    errDir="y", errBarType="both",
                    errValType="cust",
                    noEndCap=False,
                    plus=NumDataSource(numRef=NumRef(f=err_rng)),
                    minus=NumDataSource(numRef=NumRef(f=err_rng)),
                )

            chart.series.append(s)

        ws_ov.add_chart(chart, anchor)

    # --- Overview table: avg signals + %FS colour + inter-group similarity ---
    _HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    _HDR_FONT  = Font(bold=True, color="FFFFFF")
    _GRP_FILL  = PatternFill("solid", fgColor="BDD7EE")
    _GRP_FONT  = Font(bold=True)
    _GOOD_FILL = PatternFill("solid", fgColor="C6EFCE")
    _WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
    _BAD_FILL  = PatternFill("solid", fgColor="FFC7CE")

    # RSD% thresholds: (SD / mean) × 100  (intra-group reproducibility)
    _RSD_GOOD = 2.0   # ≤ 2 % → green
    _RSD_WARN = 5.0   # ≤ 5 % → yellow   > 5 % → red

    def _rsd_fill(v):
        return _GOOD_FILL if v <= _RSD_GOOD else _WARN_FILL if v <= _RSD_WARN else _BAD_FILL

    # p-value thresholds (same for all three stat tests)
    _P_GOOD = 0.05   # p > 0.05 → green  (H0 not rejected — curves match)
    _P_WARN = 0.01   # p > 0.01 → yellow   p ≤ 0.01 → red

    def _p_fill(p):
        if p is None: return PatternFill()
        return _GOOD_FILL if p > _P_GOOD else _WARN_FILL if p > _P_WARN else _BAD_FILL

    _THIN      = Side(style="thin")
    _BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    _CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)

    STANDARDS    = [f"YP50-{i}" for i in range(7)]
    ALL_ELEMENTS = list(ELEMENT_CONFIG) + EXTRA_ELEMENTS

    # Column layout:
    #   A = Element
    #   For each standard (7): n_groups × (signal col + RSD% col)
    COLS_PER_STD = n_groups * 2   # signal + RSD% per group, interleaved

    def _avg_col(std_i, si): return 2 + std_i * COLS_PER_STD + si * 2
    def _rsd_col(std_i, si): return 2 + std_i * COLS_PER_STD + si * 2 + 1

    ws_tbl      = wb.create_sheet("Overview", 0)
    prec_col    = 2 + COLS_PER_STD * len(STANDARDS)      # Precision F / Paired t
    slope_col   = prec_col + 1                            # Slope (ANCOVA)
    icpt_col    = prec_col + 2                            # Intercept (ANCOVA)
    title_end   = icpt_col

    _RSD_FILL = PatternFill("solid", fgColor="D6E4F0")

    # Row 1: title
    ws_tbl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end)
    tc = ws_tbl.cell(1, 1)
    tc.value     = "XRF Calibration Signal Overview — " + ", ".join(set_labels)
    tc.font      = Font(bold=True, size=13, color="FFFFFF")
    tc.fill      = _HDR_FILL
    tc.alignment = _CENTER
    ws_tbl.row_dimensions[1].height = 22

    # Row 2: standard group headers (each spans COLS_PER_STD cols)
    c = ws_tbl.cell(2, 1)
    c.value, c.font, c.fill, c.alignment, c.border = "Element", _HDR_FONT, _HDR_FILL, _CENTER, _BORDER
    for std_i, std in enumerate(STANDARDS):
        col_s = _avg_col(std_i, 0)
        col_e = col_s + COLS_PER_STD - 1
        ws_tbl.merge_cells(start_row=2, start_column=col_s, end_row=2, end_column=col_e)
        c = ws_tbl.cell(2, col_s)
        c.value, c.font, c.fill, c.alignment, c.border = std, _GRP_FONT, _GRP_FILL, _CENTER, _BORDER
    # Merge the three stat columns under a single "Statistics" banner
    ws_tbl.merge_cells(start_row=2, start_column=prec_col, end_row=2, end_column=icpt_col)
    c = ws_tbl.cell(2, prec_col)
    c.value, c.font, c.fill, c.alignment, c.border = "Statistics (2-group)", _HDR_FONT, _HDR_FILL, _CENTER, _BORDER
    ws_tbl.row_dimensions[2].height = 18

    # Row 3: per-group signal + RSD% sub-headers
    ws_tbl.cell(3, 1).border = _BORDER
    for std_i in range(len(STANDARDS)):
        for si, (lbl, _, _, _) in enumerate(group_data):
            c = ws_tbl.cell(3, _avg_col(std_i, si))
            c.value, c.font, c.fill, c.alignment, c.border = (
                lbl, Font(bold=True, size=8), _GRP_FILL, _CENTER, _BORDER
            )
            c = ws_tbl.cell(3, _rsd_col(std_i, si))
            c.value, c.font, c.fill, c.alignment, c.border = (
                "RSD%", Font(italic=True, size=9), _RSD_FILL, _CENTER, _BORDER
            )
    for _col, _lbl in [
        (prec_col,  "Precision\n(F or Paired t)"),
        (slope_col, "Slope\n(ANCOVA)"),
        (icpt_col,  "Intercept\n(ANCOVA)"),
    ]:
        c = ws_tbl.cell(3, _col)
        c.value, c.font, c.fill, c.alignment, c.border = (
            _lbl, Font(italic=True, size=9), _GRP_FILL, _CENTER, _BORDER
        )
    ws_tbl.row_dimensions[3].height = 36

    # Data rows
    _noise_results = {}   # elem -> {p, rsd_a, rsd_b}
    for row_i, elem in enumerate(ALL_ELEMENTS, start=4):
        c = ws_tbl.cell(row_i, 1)
        c.value, c.font, c.alignment, c.border = elem, Font(bold=True), _CENTER, _BORDER

        for std_i, std in enumerate(STANDARDS):
            for si, (lbl, mean_cal, sd_cal, n) in enumerate(group_data):
                # Signal cell
                avg = mean_cal.get(std, {}).get(elem)
                c   = ws_tbl.cell(row_i, _avg_col(std_i, si))
                c.alignment, c.border, c.number_format = _CENTER, _BORDER, "0.0"
                if avg is not None:
                    c.value = round(avg, 1)

                # RSD% cell
                cr = ws_tbl.cell(row_i, _rsd_col(std_i, si))
                cr.alignment, cr.border = _CENTER, _BORDER
                sd = sd_cal.get(std, {}).get(elem)
                if n >= 2 and sd is not None and avg is not None and abs(avg) >= 1.0:
                    rsd = (sd / abs(avg)) * 100
                    cr.value         = round(rsd, 1)
                    cr.number_format = '0.0"%"'
                    cr.fill          = _rsd_fill(rsd)
                else:
                    cr.value = "—"
                    cr.fill  = _RSD_FILL

        # --- Three stat columns (2-group only) ---
        for _col in (prec_col, slope_col, icpt_col):
            c = ws_tbl.cell(row_i, _col)
            c.alignment, c.border = _CENTER, _BORDER

        if n_groups == 2:
            lbl_a, mean_a, sd_a, n_a = group_data[0]
            lbl_b, mean_b, sd_b, n_b = group_data[1]

            ys_a = [mean_a.get(s, {}).get(elem) for s in STANDARDS]
            ys_b = [mean_b.get(s, {}).get(elem) for s in STANDARDS]

            # True concentrations — zero for uncalibrated/extra elements
            xs_full = [YP50_TRUE_CONC[i].get(elem, 0.0) for i in range(7)]
            has_conc = any(x > 0 for x in xs_full)

            # Keep only standards where both groups have a value
            pairs = [(x, a, b) for x, a, b in zip(xs_full, ys_a, ys_b)
                     if a is not None and b is not None]
            xs_p  = [p[0] for p in pairs]
            ys_ap = [p[1] for p in pairs]
            ys_bp = [p[2] for p in pairs]

            # 1. Precision: F-test (both n>=2) or paired t-test fallback
            cp = ws_tbl.cell(row_i, prec_col)
            if n_a >= 2 and n_b >= 2:
                f_stat, p_val = _precision_ftest(sd_a, n_a, sd_b, n_b, elem, STANDARDS)
                if f_stat is not None:
                    label = f"F={f_stat:.2f}\np={p_val:.3f}" if p_val is not None else f"F={f_stat:.2f}"
                    cp.value = label
                    cp.fill  = _p_fill(p_val)
                else:
                    cp.value = "—"
            else:
                t_stat, p_val = _paired_ttest(ys_ap, ys_bp)
                if t_stat is not None:
                    label = f"t={t_stat:.2f}\np={p_val:.3f}" if p_val is not None else f"t={t_stat:.2f}"
                    cp.value = label + "\n(paired)"
                    cp.fill  = _p_fill(p_val)
                else:
                    cp.value = "—"

            _prec_p = p_val

            # 2. Slope (ANCOVA) — N/A for uncalibrated elements
            cs2 = ws_tbl.cell(row_i, slope_col)
            if has_conc:
                t_stat, p_val = _compare_slopes(xs_p, ys_ap, ys_bp)
                if t_stat is not None:
                    label = f"t={t_stat:.2f}\np={p_val:.3f}" if p_val is not None else f"t={t_stat:.2f}"
                    cs2.value = label
                    cs2.fill  = _p_fill(p_val)
                else:
                    cs2.value = "—"
            else:
                cs2.value = "N/A\n(no ref)"

            # 3. Intercept (ANCOVA) for calibrated; Welch t-test on means for uncalibrated
            ci = ws_tbl.cell(row_i, icpt_col)
            if has_conc:
                t_stat, p_val = _compare_intercepts(xs_p, ys_ap, ys_bp)
            else:
                t_stat, p_val = _welch_ttest(ys_ap, ys_bp)
            if t_stat is not None:
                label = f"t={t_stat:.2f}\np={p_val:.3f}" if p_val is not None else f"t={t_stat:.2f}"
                ci.value = label + ("\n(mean t)" if not has_conc else "")
                ci.fill  = _p_fill(p_val)
            else:
                ci.value = "—"
            # Avg RSD per group for noise summary
            def _avg_rsd_grp(mc, sc, n_):
                vals = [sc.get(s, {}).get(elem) / abs(mc.get(s, {}).get(elem)) * 100
                        for s in STANDARDS
                        if n_ >= 2
                        and mc.get(s, {}).get(elem) is not None
                        and sc.get(s, {}).get(elem) is not None
                        and abs(mc.get(s, {}).get(elem)) >= 1.0]
                return sum(vals) / len(vals) if vals else None
            _noise_results[elem] = {
                'p':     _prec_p,
                'rsd_a': _avg_rsd_grp(mean_a, sd_a, n_a),
                'rsd_b': _avg_rsd_grp(mean_b, sd_b, n_b),
            }
        else:
            for _col in (prec_col, slope_col, icpt_col):
                ws_tbl.cell(row_i, _col).value = "N/A\n(k>2)"

    # --- Color legend ---
    _LEG_HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
    _LEG_LBL_FILL = PatternFill("solid", fgColor="EEF2F8")
    _THIN2        = Side(style="thin")
    _LEG_BORDER   = Border(left=_THIN2, right=_THIN2, top=_THIN2, bottom=_THIN2)
    _LEFT         = Alignment(horizontal="left", vertical="center")

    _data_end = 4 + len(ALL_ELEMENTS) - 1   # last data row
    summ_row  = _data_end + 2                  # one blank row gap
    leg_row   = summ_row + 4                   # title + content + blank

    # --- Precision noise summary ---
    if n_groups == 2 and _noise_results:
        lbl_a = group_data[0][0]
        lbl_b = group_data[1][0]
        a_noisier = [e for e, r in _noise_results.items()
                     if r['p'] is not None and r['p'] <= _P_GOOD
                     and r['rsd_a'] is not None and r['rsd_b'] is not None
                     and r['rsd_a'] > r['rsd_b']]
        b_noisier = [e for e, r in _noise_results.items()
                     if r['p'] is not None and r['p'] <= _P_GOOD
                     and r['rsd_a'] is not None and r['rsd_b'] is not None
                     and r['rsd_b'] > r['rsd_a']]
        no_diff   = [e for e, r in _noise_results.items()
                     if r['p'] is None or r['p'] > _P_GOOD]

        def _fmt(elems): return ", ".join(elems) if elems else "none"

        summ_text = (
            f"{lbl_a} is noisier (higher avg RSD):  {_fmt(a_noisier)}\n"
            f"{lbl_b} is noisier (higher avg RSD):  {_fmt(b_noisier)}\n"
            f"No significant variance difference (p > {_P_GOOD:g}):  {_fmt(no_diff)}"
        )

        # Title
        _SUMM_FILL = PatternFill("solid", fgColor="BDD7EE")
        ws_tbl.merge_cells(start_row=summ_row, start_column=1,
                           end_row=summ_row, end_column=10)
        sc = ws_tbl.cell(summ_row, 1)
        sc.value     = "Precision Variability Summary  (F-test p \u2264 0.05)"
        sc.font      = Font(bold=True, size=10)
        sc.fill      = _SUMM_FILL
        sc.alignment = _CENTER
        sc.border    = _LEG_BORDER
        ws_tbl.row_dimensions[summ_row].height = 16

        # Content
        ws_tbl.merge_cells(start_row=summ_row + 1, start_column=1,
                           end_row=summ_row + 1, end_column=10)
        sc2 = ws_tbl.cell(summ_row + 1, 1)
        sc2.value     = summ_text
        sc2.font      = Font(size=9)
        sc2.fill      = _LEG_LBL_FILL
        sc2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sc2.border    = _LEG_BORDER
        ws_tbl.row_dimensions[summ_row + 1].height = 48

    # Legend uses 4 columns: Context | Green | Yellow | Red
    # Each colour column carries both the fill and the descriptive text.
    _legend_rows = [
        (
            "RSD%  =  (SD ÷ mean) × 100 per group per standard  (n≥2 only)",
            _GOOD_FILL, f"RSD ≤ {_RSD_GOOD:g}%  — Excellent precision",
            _WARN_FILL, f"RSD {_RSD_GOOD:g}–{_RSD_WARN:g}%  — Acceptable",
            _BAD_FILL,  f"RSD > {_RSD_WARN:g}%  — Poor precision",
        ),
        (
            "Precision  =  F-test on pooled within-run MSE (both n≥2); else paired t-test",
            _GOOD_FILL, f"p > {_P_GOOD:g}  — Equal precision",
            _WARN_FILL, f"p {_P_WARN:g}–{_P_GOOD:g}  — Marginal",
            _BAD_FILL,  f"p ≤ {_P_WARN:g}  — Unequal precision",
        ),
        (
            "Slope (ANCOVA)  =  t-test on OLS slope difference  (H₀: sensitivities equal)",
            _GOOD_FILL, f"p > {_P_GOOD:g}  — Slopes equivalent",
            _WARN_FILL, f"p {_P_WARN:g}–{_P_GOOD:g}  — Marginal",
            _BAD_FILL,  f"p ≤ {_P_WARN:g}  — Sensitivity differs",
        ),
        (
            "Intercept (ANCOVA)  =  t-test on OLS intercept difference  (H₀: backgrounds equal)",
            _GOOD_FILL, f"p > {_P_GOOD:g}  — Backgrounds equivalent",
            _WARN_FILL, f"p {_P_WARN:g}–{_P_GOOD:g}  — Marginal",
            _BAD_FILL,  f"p ≤ {_P_WARN:g}  — Background differs",
        ),
    ]

    # --- Auto-fit column widths (data table only; legend uses merge_cells) ---
    ws_tbl.column_dimensions["A"].width = 10

    # Signal columns: fit to group label; RSD% cols narrow
    for std_i in range(len(STANDARDS)):
        for si, (lbl, _, _, _) in enumerate(group_data):
            w = max(len(lbl.split("\n")[0]) + 1, 10)
            ws_tbl.column_dimensions[get_column_letter(_avg_col(std_i, si))].width = w
            ws_tbl.column_dimensions[get_column_letter(_rsd_col(std_i, si))].width = 6

    # Stat columns
    for _col in (prec_col, slope_col, icpt_col):
        ws_tbl.column_dimensions[get_column_letter(_col)].width = 13

    # --- Color legend (fixed-width spans, independent of data table) ---
    # Context = 4 cols wide; each colour band = 2 cols wide  (total = 10)
    _leg_spans = [
        (1, 4),   # Context
        (5, 6),   # Green
        (7, 8),   # Yellow
        (9, 10),  # Red
    ]

    def _leg_cell(row, cs, ce, value, fill, font, align):
        if ce > cs:
            ws_tbl.merge_cells(start_row=row, start_column=cs,
                               end_row=row,   end_column=ce)
        c = ws_tbl.cell(row, cs)
        c.value, c.font, c.fill, c.alignment, c.border = (
            value, font, fill, align, _LEG_BORDER
        )

    # Title row
    _leg_cell(leg_row, 1, 10,
              "Color Code Legend",
              _LEG_HDR_FILL, Font(bold=True, size=11), _CENTER)
    ws_tbl.row_dimensions[leg_row].height = 18

    # Sub-header row
    for (cs, ce), hdr in zip(_leg_spans,
                              ["Context",
                               "✓ Match (green)",
                               "⚠ Marginal (yellow)",
                               "✗ Differs (red)"]):
        _leg_cell(leg_row + 1, cs, ce, hdr,
                  _LEG_HDR_FILL, Font(bold=True, size=9), _CENTER)

    # Data rows
    _lft_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r_off, (ctx, gf, gt, yf, yt, rf, rt) in enumerate(_legend_rows, start=2):
        row = leg_row + r_off
        _leg_cell(row, *_leg_spans[0], ctx,  _LEG_LBL_FILL, Font(size=9), _lft_wrap)
        _leg_cell(row, *_leg_spans[1], gt,   gf,             Font(size=9), _LEFT)
        _leg_cell(row, *_leg_spans[2], yt,   yf,             Font(size=9), _LEFT)
        _leg_cell(row, *_leg_spans[3], rt,   rf,             Font(size=9), _LEFT)
        ws_tbl.row_dimensions[row].height = 16

    # --- Raw individual sheets for multi-file groups (appended at the end) ---
    raw_sheet_labels = []
    ws_cal_template  = wb[set_labels[0]]   # use first averaged sheet as copy source
    raw_seen: set    = set(seen_labels)

    for si, (paths, (label, mean_cal, sd_cal, n)) in enumerate(zip(group_paths, group_data)):
        if n < 2:
            continue
        for p in paths:
            parsed   = parse_xrf_txt(p)
            cal_sets = parsed.get("cal_sets", [])
            if not cal_sets:
                continue
            rd = parsed.get("run_date") or ""
            for set_i, cal_set in enumerate(cal_sets):
                suffix    = f" #{set_i + 1}" if len(cal_sets) > 1 else ""
                base_max  = 28 - len(suffix)
                raw_base  = f"{rd or p.stem} ({p.stem})"[:base_max] + suffix
                raw_label = _unique_label(raw_base, p.stem, raw_seen)

                ws_raw = wb.copy_worksheet(ws_cal_template)
                ws_raw.title = raw_label

                for std_idx, elem_col, sig_col in _CALDAT_COLS:
                    std_key = f"YP50-{std_idx}"
                    signals = cal_set.get(std_key, {})
                    for row in range(2, ws_raw.max_row + 1):
                        ev = ws_raw.cell(row, elem_col).value
                        if ev and isinstance(ev, str) and ev.strip() in signals:
                            ws_raw.cell(row, sig_col, signals[ev.strip()])

                _autosize_columns(ws_raw)
                raw_sheet_labels.append(raw_label)
                print(f"  Raw sheet '{raw_label}' ← {p.name}{suffix}")

    # Sheet order: overview + averaged Caldats + elements + raw individual sheets
    desired_order = (
        ["Overview", "Overview Graphs"]
        + set_labels
        + list(ELEMENT_CONFIG)
        + raw_sheet_labels
    )
    wb._sheets = [wb[name] for name in desired_order]

    wb.save(output)
    print(f"\nComparison report written: {output}")
    print(f"  Sheets: Overview | Overview Graphs | {' | '.join(set_labels)} | <element sheets>")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Compare N XRF calibration sets. "
                    "Bracket-quoted groups (e.g. \"[a.txt,b.txt]\") are treated as "
                    "repeated measurements and averaged with SD error bars."
    )
    parser.add_argument(
        "files", nargs="+",
        help="XRF .txt files or bracket-quoted groups, e.g. \"[a.txt,b.txt]\".",
    )
    parser.add_argument(
        "--output", default=None,
        help="Output .xlsx path.  Defaults to calset_comparison_<stems>.xlsx.",
    )
    args = parser.parse_args()

    groups = parse_file_groups(args.files)

    # Validate all paths exist
    for grp in groups:
        for p in grp:
            if not p.exists():
                print(f"Error: file not found: {p}")
                sys.exit(1)

    if len(groups) < 2:
        print("Error: provide at least two files/groups to compare.")
        sys.exit(1)

    if args.output:
        out = pathlib.Path(args.output)
    else:
        first_stems = [grp[0].stem for grp in groups[:3]]
        stems = "_vs_".join(first_stems)
        if len(groups) > 3:
            stems += f"_and_{len(groups) - 3}_more"
        out = groups[0][0].parent / f"calset_comparison_{stems}.xlsx"

    print(f"Building calset comparison for {len(groups)} group(s)...")
    build_comparison(groups, out)


if __name__ == "__main__":
    main()

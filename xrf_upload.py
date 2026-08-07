"""
XRF Uploader — core logic
Parses .txt instrument files, applies two-range linear calibration,
generates per-sample .xlsx reports, then uploads Benchling Copyable
Tables to three assay schemas.

NOTE: Benchling field names below are best-guess from schema IDs.
      Verify/update FIELD_* constants against your actual schemas.
"""

import csv
import datetime
import pathlib
import re
import shutil

import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import PatternFill
from benchling_sdk.models import AssayResultCreate
from benchling_sdk.helpers.serialization_helpers import fields as _fields

# ---------------------------------------------------------------------------
# Schema IDs
# ---------------------------------------------------------------------------
XRF_CONCENTRATION_SCHEMA = "assaysch_e7BnMa7Z" 
XRF_SIGNAL_SCHEMA        = "assaysch_dHcQUAwe"
XRF_SUMMARY_SCHEMA       = "assaysch_9LFkZwZL"

# Benchling URL prefix (no trailing slash)
BENCHLING_URL = "https://florrent.benchling.com"

# R&D Recipe Project → 4-Metrology → X-Ray Spectroscopy
XRF_PROJECT_ID = "src_9m7pHy8Q"
XRF_FOLDER_ID  = "lib_ztwU5Vd1"

# Maps name prefix (as it appears in .txt Ident column) → Benchling entity schema ID.
# Benchling entity names have lineage appended (e.g. "Char-163-[Biomass-...]"),
# so we match by prefix within the correct schema.
ENTITY_SCHEMA_MAP = {
    "Char":              "ts_8U1FBPE1",
    "Biomass":           "ts_lr0RcaJe",
    "Ash":               "ts_Ydz3BUer4L",
    "Contoured_Carbon":  "ts_6xF4VfBV",
    "Annealate":         "ts_CyHUVI6V",
    "Ann":               "ts_CyHUVI6V",  # abbreviated form used in instrument
    "Pyrolite":          "ts_MGUIIAjD",
    "Millate":           "ts_LB1Um3kj",
    "Washate":           "ts_1xh8kqjP",
    "Chemical_Mixture":  "ts_iG1bjoQg",  # must come before "Chemical"
    "Chemical":          "ts_yw8qNvZZ",
    "Slurry":            "ts_E2wVbIPI",
    "Electrode_Slurry":  "ts_KV5l0zvA",
    "Solution":          "ts_E1pPkSJ0",
    "Coating":           "ts_fSepyueW",
    "Hemp":              "ts_9bxPgu32",
    "Preactivate":       "ts_s4Lv4JFO",
    "Activate":          "ts_HIk0EDoe",
    "Electrolyte":       "ts_nOzTUoUE",
    "Binder_Dispersion": "ts_KAWds1xZ",
    "Cell_Component":    "ts_ZcaHYk9c",
    "Crucible":          "ts_AcIYpJjd",
}

# Abbreviation expansions: instrument Ident prefix → full Benchling entity prefix.
# Used when the .txt file uses a shortened name that differs from the Benchling name.
ENTITY_NAME_EXPANSIONS = {
    "Ann-": "Annealate-",
}

# ---------------------------------------------------------------------------
# Dropdown option ID maps
# ---------------------------------------------------------------------------
# XRF Calibrated Elements (sfs_frC6BvrU) — used in concentration schema
CONC_ELEMENT_IDS = {
    "Potassium (K)":   "sfso_pv4YLSFY",
    "Calcium (Ca)":    "sfso_UHH3TSu2",
    "Vanadium (V)":    "sfso_B9eIjO3g",
    "Chromium (Cr)":   "sfso_Syd8Vcfn",
    "Manganese (Mn)":  "sfso_Ns6aq74s",
    "Iron (Fe)":       "sfso_Xu62n82o",
    "Cobalt (Co)":     "sfso_WbKBIbjR",
    "Nickel (Ni)":     "sfso_vejLGAny",
    "Copper (Cu)":     "sfso_5htMuXS3",
    "Zinc (Zn)":       "sfso_hHqRu4iZ",
    "Gallium (Ga)":    "sfso_N6DTkRUO",
    "Rubidium (Rb)":   "sfso_7YW9SRco",
    "Strontium (Sr)":  "sfso_55YPu8w6",
    "Molybdenum (Mo)": "sfso_lmZqXVss",
    "Cadmium (Cd)":    "sfso_9xC9zjFA",
    "Thallium (Tl)":   "sfso_uq7ot3IX",
    "Lead (Pb)":       "sfso_3n1GRuSP",
    "Bismuth (Bi)":    "sfso_pg4kmRvy",
}

# XRF Elements (sfs_CLPJT0Nt) — used in signal intensity schema
SIG_ELEMENT_IDS = {
    "Potassium (K)":   "sfso_4MQjZ7Kn",
    "Calcium (Ca)":    "sfso_iEKBtSH8",
    "Vanadium (V)":    "sfso_FCxENzYg",
    "Chromium (Cr)":   "sfso_oQYhnwcV",
    "Manganese (Mn)":  "sfso_KYHboSE1",
    "Iron (Fe)":       "sfso_VnyKh88M",
    "Cobalt (Co)":     "sfso_cKPmtoej",
    "Nickel (Ni)":     "sfso_4P7j7JuF",
    "Copper (Cu)":     "sfso_bhCzOopm",
    "Zinc (Zn)":       "sfso_3xSVo4P0",
    "Gallium (Ga)":    "sfso_ZruD5ltd",
    "Rubidium (Rb)":   "sfso_sM0SLj7S",
    "Strontium (Sr)":  "sfso_7GiZxVme",
    "Molybdenum (Mo)": "sfso_qTOpcRad",
    "Cadmium (Cd)":    "sfso_laYUi87K",
    "Thallium (Tl)":   "sfso_s5snuAgf",
    "Lead (Pb)":       "sfso_x7L0uJcS",
    "Bismuth (Bi)":    "sfso_959Xj7Zh",
    "Arsenic (As)":    "sfso_BzmxHYJG",
    "Cerium (Ce)":     "sfso_A8msucVi",
    "Cesium (Cs)":     "sfso_NWrJfMEo",
    "Chlorine (Cl)":   "sfso_kbzSUzfW",
    "Iridium (Ir)":    "sfso_ICRk5hfy",
    "Phosphorus (P)":  "sfso_e51gK0Xh",
    "Silicon (Si)":    "sfso_P7AortlC",
    "Sulfur (S)":      "sfso_9b7ORs37",
    "Tellurium (Te)":  "sfso_M2VBpky5",
    "Tin (Sn)":        "sfso_zHa4zEvU",
    "Titanium (Ti)":   "sfso_diuDbjIF",
    "Tungsten (W)":    "sfso_CxPm4AfZ",
    "Yttrium (Y)":     "sfso_vO7n93LO",
    "Zirconium (Zr)":  "sfso_qFhbLxAS",
}

# XRF Signal Quality (sfs_lQA8RmQA)
SIGNAL_QUALITY_IDS = {
    "In Range":        "sfso_NNxxw7nl",
    "High Signal":     "sfso_r6ebAifK",
    "Low Signal":      "sfso_ihGI4xfV",
    "Not Significant": "sfso_U1HgeMxn",
}

# ---------------------------------------------------------------------------
# Field name constants — update these to match your Benchling schema fields
# ---------------------------------------------------------------------------
# Concentration schema fields
FIELD_CONC_SAMPLE   = "material" #tsf_lxSJeyHJ
FIELD_CONC_ELEMENT  = "element" #tsf_SFX7VDa4
FIELD_CONC_VALUE    = "concentration_ppm" #tsf_LA6zIgyH
FIELD_CONC_QUALITY  = "signal_quality" #tsf_L8Y9TNIg
#FIELD_CONC_LOD      = "lod"

# Signal intensity schema fields
FIELD_SIG_SAMPLE    = "material" #tsf_nKDnwU2L
FIELD_SIG_ELEMENT   = "element" #tsf_Dv2603LU
FIELD_SIG_VALUE     = "signal_intensity" #tsf_4vwG3S2p

# Summary schema fields
FIELD_SUM_SAMPLE    = "material" #tsf_NYZy54Gv
FIELD_SUM_REPORT    = "xrf_analysis_report_xlsx" #tsf_TL2VjurB
#FIELD_SUM_DATE      = "date"

# ---------------------------------------------------------------------------
# Calibrated-element configuration
# LOD in ppm, range indices into [YP50-0 .. YP50-6]
# excl_idx = index excluded from range fitting (background standard)
# ---------------------------------------------------------------------------
ELEMENT_CONFIG = {
    #          lod   r1_idx     r2_idx       excl_idx
    "K":   dict(lod=15,  r1=[0,1,2,3], r2=[4,5,6], excl=[]),
    "Ca":  dict(lod=15,  r1=[0,1,2],   r2=[3,4,5,6], excl=[]),
    "V":   dict(lod=10,  r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Cr":  dict(lod=8,   r1=[0,1,2,3], r2=[4,5,6], excl=[]),
    "Mn":  dict(lod=6,   r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Fe":  dict(lod=5,   r1=[0,1,2,3], r2=[4,5,6], excl=[]),
    "Co":  dict(lod=5,   r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Ni":  dict(lod=4,   r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Cu":  dict(lod=4,   r1=[0,1,2],   r2=[3,4,5,6], excl=[]),
    "Zn":  dict(lod=5,   r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Ga":  dict(lod=5,   r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Rb":  dict(lod=4,   r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Sr":  dict(lod=3,   r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Mo":  dict(lod=3,   r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Cd":  dict(lod=15,  r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Tl":  dict(lod=10,  r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Pb":  dict(lod=10,  r1=[1,2,3],   r2=[4,5,6], excl=[0]),
    "Bi":  dict(lod=10,  r1=[1,2,3],   r2=[4,5,6], excl=[0]),
}

ELEMENT_FULL_NAMES = {
    "K": "Potassium (K)", "Ca": "Calcium (Ca)", "V": "Vanadium (V)",
    "Cr": "Chromium (Cr)", "Mn": "Manganese (Mn)", "Fe": "Iron (Fe)",
    "Co": "Cobalt (Co)", "Ni": "Nickel (Ni)", "Cu": "Copper (Cu)",
    "Zn": "Zinc (Zn)", "Ga": "Gallium (Ga)", "Rb": "Rubidium (Rb)",
    "Sr": "Strontium (Sr)", "Mo": "Molybdenum (Mo)", "Cd": "Cadmium (Cd)",
    "Tl": "Thallium (Tl)", "Pb": "Lead (Pb)", "Bi": "Bismuth (Bi)",
    # Uncalibrated / monitor elements
    "Al": "Aluminum (Al)",   "Am": "Americium (Am)", "As": "Arsenic (As)",
    "Br": "Bromine (Br)",    "Ce": "Cerium (Ce)",    "Cl": "Chlorine (Cl)",
    "Er": "Erbium (Er)",     "Eu": "Europium (Eu)",  "Hf": "Hafnium (Hf)",
    "In": "Indium (In)",     "Ir": "Iridium (Ir)",   "La": "Lanthanum (La)",
    "Nb": "Niobium (Nb)",    "Nd": "Neodymium (Nd)", "Os": "Osmium (Os)",
    "P":  "Phosphorus (P)",  "Pt": "Platinum (Pt)",  "Re": "Rhenium (Re)",
    "Rh": "Rhodium (Rh)",    "S":  "Sulfur (S)",      "Si": "Silicon (Si)",
    "Sn": "Tin (Sn)",        "Ta": "Tantalum (Ta)",   "Te": "Tellurium (Te)",
    "Th": "Thorium (Th)",    "Ti": "Titanium (Ti)",   "Y":  "Yttrium (Y)",
    "Yb": "Ytterbium (Yb)",  "Zr": "Zirconium (Zr)",
}

# Elements measured by the instrument but without YP50 calibration reference values.
# Shown in the comparison overview (signals + RSD + precision F-test + mean t-test)
# but excluded from calibration curve building and sample concentration calculation.
EXTRA_ELEMENTS = [
    "Al", "Am", "As", "Br", "Ce", "Cl", "Er", "Eu", "Hf",
    "In", "Ir", "La", "Nb", "Nd", "Os", "P",  "Pt", "Re",
    "Rh", "S",  "Si", "Sn", "Ta", "Te", "Th", "Ti", "Y",
    "Yb", "Zr",
]

# True concentrations for YP50-0..6, ordered by ELEMENT_CONFIG keys
_YP50_CONC_ROWS = [
    # K        Ca            V         Cr        Mn        Fe         Co           Ni       Cu        Zn        Ga        Rb        Sr        Mo        Cd        Tl        Pb        Bi
    (125.146,  206.2861,     0,        6.4965,   0,        34.79,     0,           0.461,   20.54,    0,        0,        0,        0,        0,        0,        0,        0,        0),
    (225.866,  307.005907,   10.07198, 16.56848, 10.07198, 135.5098,  10.0719807,  10.53298,30.61198, 10.07198, 10.07198, 10.07198, 10.07198, 10.07198, 10.07198, 10.07198, 10.07198, 10.0719807),
    (308.777,  389.916665,   18.36306, 24.85956, 18.36306, 218.4206,  18.3630565,  18.82406,38.90306, 18.36306, 18.36306, 18.36306, 18.36306, 18.36306, 18.36306, 18.36306, 18.36306, 18.3630565),
    (491.25,   572.3897,     36.61036, 43.10686, 36.61036, 400.8936,  36.61036,    37.07136,57.15036, 36.61036, 36.61036, 36.61036, 36.61036, 36.61036, 36.61036, 36.61036, 36.61036, 36.61036),
    (860.159,  941.299211,   73.50131, 79.99781, 73.50131, 769.8031,  73.5013111,  73.96231,94.04131, 73.50131, 73.50131, 73.50131, 73.50131, 73.50131, 73.50131, 73.50131, 73.50131, 73.5013111),
    (1214.036, 1295.17641,   108.889,  115.3855, 108.889,  1123.68,   108.889031,  109.35,  129.429,  108.889,  108.889,  108.889,  108.889,  108.889,  108.889,  108.889,  108.889,  108.889031),
    (1578.199, 1659.33943,   145.3053, 151.8018, 145.3053, 1487.843,  145.305333,  145.7663,165.8453, 145.3053, 145.3053, 145.3053, 145.3053, 145.3053, 145.3053, 145.3053, 145.3053, 145.305333),
]
_CAL_ELEMS = list(ELEMENT_CONFIG.keys())
YP50_TRUE_CONC = {
    i: dict(zip(_CAL_ELEMS, row)) for i, row in enumerate(_YP50_CONC_ROWS)
}

CALIBRATION_STANDARDS = [f"YP50F-{i}" for i in range(7)]  # Ident prefixes in txt

# ---------------------------------------------------------------------------
# 1. TXT PARSER
# ---------------------------------------------------------------------------

def parse_xrf_txt(txt_path: pathlib.Path) -> dict:
    """
    Parse an XRF instrument .txt file.

    Returns:
        {
          'samples':        {ident: {element: signal_float}},
          'cal_standards':  {std_name: {element: signal_float}},  # first occurrence only
          'all_elements':   [ordered list of all elements in file],
          'run_date':       str or None,
        }
    """
    txt_path = pathlib.Path(txt_path)
    with txt_path.open(encoding="utf-8", errors="replace") as fh:
        content = fh.read()

    # Detect delimiter (tab or semicolon)
    delim = "\t" if "\t" in content[:500] else ";"
    reader = csv.reader(content.splitlines(), delimiter=delim)
    rows = list(reader)

    # Row 0: element symbols (padded with spaces)
    # Row 1: "Nr", "Ident", "Seq", "Time", "Pos", "C", "Unit", "C", "Unit", ...
    # Locate the Nr/Ident/Seq header row, then step back one to get the element row.
    elem_header_row = None
    for i, row in enumerate(rows):
        stripped = [c.strip() for c in row]
        if "Nr" in stripped and "Ident" in stripped and "Seq" in stripped:
            if i > 0:
                elem_header_row = i - 1
            break
    if elem_header_row is None:
        raise ValueError("Cannot find element header row in txt file")

    header = [c.strip() for c in rows[elem_header_row]]

    # The data header row (row below) has Nr, Ident, Seq, Time, Pos, C, Unit, C, Unit...
    # Element columns alternate: symbol_col → data_col in data rows, unit_col = data_col+1
    # We map each element to the column index it occupies in the ELEMENT row,
    # which is the same column as the value ("C") in data rows.
    elem_cols = {}
    i = 0
    while i < len(header):
        cell = header[i]
        if cell and cell not in ("Nr", "Ident", "Seq", "Time", "Pos", ""):
            if cell not in elem_cols:
                elem_cols[cell] = i
            i += 2  # skip unit column
        else:
            i += 1

    all_elements = list(elem_cols.keys())

    def _sig(row, elem):
        col = elem_cols.get(elem)
        if col is None or col >= len(row):
            return 0.0
        val = row[col].strip().replace("ppm", "").strip()
        try:
            return float(val)
        except (ValueError, AttributeError):
            return 0.0

    samples = {}
    cal_sets = []
    current_set = {}
    current_set_stds = set()
    run_date = None

    for row in rows[elem_header_row + 2:]:  # skip element-symbol row AND Nr/C/Unit row
        if len(row) < 3:
            continue
        ident = row[1].strip() if len(row) > 1 else ""
        seq   = row[2].strip() if len(row) > 2 else ""
        time  = row[3].strip() if len(row) > 3 else ""
        if not ident or ident in ("Ident", "Average", "S.Dev.", "SDev", ""):
            continue
        if any(x in seq for x in ("Ave of", "SDev of", "Ave.", "S.Dev.")):
            continue
        if run_date is None and time:
            run_date = time.split()[0] if " " in time else time[:10]

        signals = {e: _sig(row, e) for e in all_elements}

        # Identify calibration standards by Ident prefix YP50F-
        is_std = any(ident.upper().startswith(s.upper()) for s in CALIBRATION_STANDARDS)
        if is_std:
            for std_i, name in enumerate(CALIBRATION_STANDARDS):
                if ident.upper().startswith(name.upper()):
                    if std_i in current_set_stds:
                        # Repeated standard index → close current set, start a new one
                        cal_sets.append(current_set)
                        current_set = {}
                        current_set_stds = set()
                    current_set[f"YP50-{std_i}"] = signals
                    current_set_stds.add(std_i)
                    if "_date" not in current_set and time:
                        current_set["_date"] = time.split()[0] if " " in time else time[:10]
                    break
        else:
            samples[ident] = signals

    if current_set:
        cal_sets.append(current_set)

    return {
        "samples":       samples,
        "cal_standards": cal_sets[0] if cal_sets else {},
        "cal_sets":      cal_sets,
        "all_elements":  all_elements,
        "run_date":      run_date,
    }

# ---------------------------------------------------------------------------
# 2. CALIBRATION ENGINE
# ---------------------------------------------------------------------------

def _linreg(xs, ys):
    """OLS linear regression y = m*x + b. Returns (m, b, r2)."""
    n = len(xs)
    if n < 2:
        return 1.0, 0.0, 0.0
    sx = sum(xs); sy = sum(ys)
    sxx = sum(x*x for x in xs)
    sxy = sum(x*y for x, y in zip(xs, ys))
    denom = n * sxx - sx * sx
    if abs(denom) < 1e-12:
        return 1.0, 0.0, 0.0
    m = (n * sxy - sx * sy) / denom
    b = (sy - m * sx) / n
    y_mean = sy / n
    ss_tot = sum((y - y_mean) ** 2 for y in ys)
    ss_res = sum((y - (m * x + b)) ** 2 for x, y in zip(xs, ys))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 1e-12 else 1.0
    return m, b, r2


def build_calibration(cal_signals: dict) -> dict:
    """
    Build two-range calibration curves for every element in ELEMENT_CONFIG.

    cal_signals: {'YP50-0': {elem: signal}, ..., 'YP50-6': {...}}
    Returns:
        {elem: {
            'r1': {'m','b','r2','min_sig','max_sig'},
            'r2': {'m','b','r2','min_sig','max_sig'},
            'r1_pts': [(conc, signal), ...],
            'r2_pts': [(conc, signal), ...],
            'excl_pts': [(conc, signal), ...],
        }}
    """
    calibration = {}
    for elem, cfg in ELEMENT_CONFIG.items():
        r1_pts = []
        r2_pts = []
        excl_pts = []
        for idx in range(7):
            sig = cal_signals.get(f"YP50-{idx}", {}).get(elem, 0.0)
            conc = YP50_TRUE_CONC[idx][elem]
            if idx in cfg["excl"]:
                excl_pts.append((conc, sig))
            elif idx in cfg["r1"]:
                r1_pts.append((conc, sig))
            elif idx in cfg["r2"]:
                r2_pts.append((conc, sig))

        def _fit(pts):
            xs = [p[0] for p in pts]
            ys = [p[1] for p in pts]
            m, b, r2 = _linreg(xs, ys)
            return {
                "m": m, "b": b, "r2": r2,
                "min_sig": min(ys), "max_sig": max(ys),
                "min_conc": min(xs), "max_conc": max(xs),
            }

        calibration[elem] = {
            "r1":      _fit(r1_pts),
            "r2":      _fit(r2_pts),
            "r1_pts":  r1_pts,
            "r2_pts":  r2_pts,
            "excl_pts": excl_pts,
        }
    return calibration


def _calc_concentration(signal: float, elem: str, cal: dict):
    """
    Calculate concentration and signal quality for one element/sample.

    Returns (conc_value_or_None, quality_string, display_conc)
    where display_conc is the value to show in the xlsx (float or '—').
    """
    cfg  = ELEMENT_CONFIG[elem]
    lod  = cfg["lod"]
    std_dev = round(lod / 3, 2)
    r1   = cal[elem]["r1"]
    r2   = cal[elem]["r2"]

    if signal is None or signal <= 0:
        return None, "No Signal", "\u2014"

    # Determine which range to use
    if signal <= r1["max_sig"]:
        fit = r1
    else:
        fit = r2  # includes above-range signals

    # Calculate concentration: conc = (signal - b) / m
    if abs(fit["m"]) < 1e-12:
        return None, "No Signal", "\u2014"
    conc = (signal - fit["b"]) / fit["m"]

    # High signal: above Range 2 max
    if signal > r2["max_sig"]:
        return conc, "High Signal", conc

    r1_min_sig = r1["min_sig"]

    # Signal below calibration range minimum
    if signal < r1_min_sig:
        if conc <= std_dev:
            return None, "Not Significant", "\u2014"
        else:
            return conc, "Low Signal", conc

    # Signal within calibration range
    if conc < lod:
        return conc, "Low Signal", conc
    return conc, "In Range", conc


def process_sample(sample_signals: dict, calibration: dict) -> dict:
    """
    Returns {elem: {'conc', 'quality', 'signal', 'display_conc', 'lod', 'std_dev'}}
    for all calibrated elements.
    """
    results = {}
    for elem in ELEMENT_CONFIG:
        sig = sample_signals.get(elem, 0.0)
        conc, quality, display = _calc_concentration(sig, elem, calibration)
        cfg = ELEMENT_CONFIG[elem]
        results[elem] = {
            "signal":      sig,
            "conc":        conc,
            "quality":     quality,
            "display_conc": display,
            "lod":         cfg["lod"],
            "std_dev":     round(cfg["lod"] / 3, 2),
        }
    return results

# ---------------------------------------------------------------------------
# 3. XLSX REPORT GENERATOR
# ---------------------------------------------------------------------------

TEMPLATE_PATH = pathlib.Path(__file__).parent / "XRF Sample Report - TEMPLATE.xlsx"

# Caldat column mapping per standard: (std_index, elem_col_1indexed, signal_col_1indexed)
_CALDAT_COLS = [
    (0,  1,  2),   # YP50-0: elem=A, signal=B
    (1,  4,  5),   # YP50-1: elem=D, signal=E
    (2,  7,  8),   # YP50-2: elem=G, signal=H
    (3, 10, 11),   # YP50-3: elem=J, signal=K
    (4, 13, 14),   # YP50-4: elem=M, signal=N
    (5, 16, 17),   # YP50-5: elem=P, signal=Q
    (6, 19, 20),   # YP50-6: elem=S, signal=T
]

_DASH = "\u2014"


def _format_run_date(run_date: str) -> str:
    """Parse run_date string (e.g. '16-Jul-2026') into YYYYMMDD. Returns 'unknown' on failure."""
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(run_date, fmt).strftime("%Y%m%d")
        except (ValueError, TypeError):
            pass
    return "unknown"


def generate_xlsx_report(
    sample_name: str,
    cal_signals: dict,
    raw_signals: dict,
    all_elements: list,
    output_dir: pathlib.Path,
    run_date: str = None,
) -> pathlib.Path:
    """
    Copy the report template and fill in sample name, raw element signals,
    and calibration standard signals. All calibration curve calculations are
    handled by the template's built-in Excel formulas.
    Returns the path to the written file.
    """
    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    date_str = _format_run_date(run_date) if run_date else "unknown"
    base_stem = f"XRF Sample Report - {sample_name} - {date_str}"
    out_path = output_dir / f"{base_stem}.xlsx"
    if out_path.exists():
        n = 2
        while (output_dir / f"{base_stem} - {n}.xlsx").exists():
            n += 1
        out_path = output_dir / f"{base_stem} - {n}.xlsx"

    shutil.copy2(TEMPLATE_PATH, out_path)
    wb = openpyxl.load_workbook(out_path)

    # --- Samdat: sample name ---
    ws_sam = wb["Samdat"]
    ws_sam["B1"] = sample_name

    # --- Samdat: raw element signals (columns A/B, rows 4+) ---
    for i, elem in enumerate(all_elements):
        ws_sam.cell(4 + i, 1, elem)
        ws_sam.cell(4 + i, 2, raw_signals.get(elem, 0.0))

    # --- Caldat: calibration standard signals ---
    ws_cal = wb["Caldat"]
    for std_idx, elem_col, sig_col in _CALDAT_COLS:
        std_key = f"YP50-{std_idx}"
        signals = cal_signals.get(std_key, {})
        for row in range(2, ws_cal.max_row + 1):
            elem = ws_cal.cell(row, elem_col).value
            if elem and isinstance(elem, str):
                elem = elem.strip()
                if elem in signals:
                    ws_cal.cell(row, sig_col, signals[elem])

    wb.save(out_path)
    return out_path


def build_benchling_rows(sample_name: str, processed: dict, raw_signals: dict) -> tuple:
    """
    Build Benchling table rows directly from Python-computed processed data.

    Returns:
        conc_rows:   [{'material','element','concentration','signal_quality','lod'}, ...]
        signal_rows: [{'material','element','signal_intensity'}, ...]
    """
    conc_rows = [
        {
            "material":       sample_name,
            "element":        ELEMENT_FULL_NAMES[elem],
            "concentration":  None if r["display_conc"] == _DASH else r["display_conc"],
            "signal_quality": r["quality"],
            "lod":            r["lod"],
        }
        for elem, r in processed.items()
    ]
    signal_rows = [
        {
            "material":         sample_name,
            "element":          ELEMENT_FULL_NAMES.get(elem, elem),
            "signal_intensity": sig,
        }
        for elem, sig in raw_signals.items()
        if sig > 0
    ]
    return conc_rows, signal_rows


def generate_xlsx_cal_report(
    cal_sets: list,
    all_elements: list,
    run_date: str,
    output_dir: pathlib.Path,
) -> pathlib.Path:
    """
    Generate a calibration-only xlsx based on the report template.

    Structure:
      - No Samdat sheet.
      - 'Caldat' sheet (Set 1) + 'Caldat 2', 'Caldat 3', ... for each extra set.
        Set 1 keeps the name 'Caldat' so all existing element-sheet formulas resolve.
      - Element sheets retain full template formatting and the 3 built-in charts;
        each additional set's data is written into spare columns and appended as new
        series on all three charts.
    """
    output_dir = pathlib.Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"XRF Calibration Report - {run_date or 'unknown'}.xlsx"

    shutil.copy2(TEMPLATE_PATH, out_path)
    wb = openpyxl.load_workbook(out_path)

    del wb["Samdat"]

    n_sets = len(cal_sets)

    # --- Build date-based set labels (needed for sheet naming) ---
    def _set_labels(sets):
        dates = [s.get("_date", "Unknown") for s in sets]
        from collections import Counter
        counts = Counter(dates)
        day_seen = {}
        labels = []
        for d in dates:
            if counts[d] == 1:
                labels.append(d)
            else:
                n = day_seen.get(d, 0) + 1
                day_seen[d] = n
                labels.append(f"{d} ({n})")
        return labels

    set_labels = _set_labels(cal_sets)

    # --- Caldat sheets: named by date label ---
    # Copy extras first while the original is still named 'Caldat', then rename it.
    ws_cal_orig = wb["Caldat"]
    caldat_sheets = [ws_cal_orig]
    for si in range(1, n_sets):
        ws_copy = wb.copy_worksheet(ws_cal_orig)
        ws_copy.title = set_labels[si]
        caldat_sheets.append(ws_copy)
    ws_cal_orig.title = set_labels[0]  # rename original last so copies are clean

    for ws_c, cal_set in zip(caldat_sheets, cal_sets):
        for std_idx, elem_col, sig_col in _CALDAT_COLS:
            std_key = f"YP50-{std_idx}"
            signals = cal_set.get(std_key, {})
            for row in range(2, ws_c.max_row + 1):
                elem = ws_c.cell(row, elem_col).value
                if elem and isinstance(elem, str):
                    elem = elem.strip()
                    if elem in signals:
                        ws_c.cell(row, sig_col, signals[elem])

    # --- Helpers ---
    def _ref_rows(formula: str):
        """Parse (min_row, max_row) from a formula like \"K!$G$2:$G$8\"."""
        m = re.search(r'\$[A-Z]+\$(\d+):\$[A-Z]+\$(\d+)', formula)
        return (int(m.group(1)), int(m.group(2))) if m else (2, 8)

    def _signal_col(si):
        """1-indexed column for Set (si+1) signal.
        si=1 → I(9), si=2 → J(10), si=3 → S(19), si=4 → T(20), ...
        Skips cols 11-18 which are the template stats table.
        """
        return 8 + si if si <= 2 else 16 + si

    def _col_letter(n):
        result = ""
        while n:
            n, r = divmod(n - 1, 26)
            result = chr(65 + r) + result
        return result

    _YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

    # --- Element sheets: add labelled signal columns + chart series for extra sets ---
    for elem in ELEMENT_CONFIG:
        ws = wb[elem]

        # Update H1 header to include Set 1 date (matches template formula style)
        ws.cell(1, 8, f'=""&$B$2&" Signal"&" - {set_labels[0]}'  + '"')

        if n_sets < 2:
            continue

        chart_ranges = [_ref_rows(ch.series[0].xVal.numRef.f) for ch in ws._charts[:3]]

        for si in range(1, n_sets):
            col_sig = _signal_col(si)

            # Header: same formula style as H1 but with this set's date
            ws.cell(1, col_sig, f'=""&$B$2&" Signal"&" - {set_labels[si]}' + '"')

            # Rows 2-8: VLOOKUP formulas copied from H column with sheet name replaced.
            # H column still says 'Caldat!' here; replace with quoted date-based name.
            for row in range(2, 9):
                h_val = ws.cell(row, 8).value
                if isinstance(h_val, str) and h_val.startswith("="):
                    ws.cell(row, col_sig).value = h_val.replace(
                        "Caldat!", f"'{set_labels[si]}'!")
                ws.cell(row, col_sig).fill = _YELLOW

            for chart_i, (rmin, rmax) in enumerate(chart_ranges):
                chart = ws._charts[chart_i]
                x_ref = Reference(ws, min_col=7, min_row=rmin, max_row=rmax)
                y_ref = Reference(ws, min_col=col_sig, min_row=rmin, max_row=rmax)
                chart.series.append(Series(y_ref, xvalues=x_ref, title=set_labels[si]))

        # Replace all remaining 'Caldat!' references in this sheet with the quoted
        # Set 1 date name (covers G col HLOOKUP, H col VLOOKUP, B5 LOD formula, etc.)
        label0 = f"'{set_labels[0]}'!"
        for ws_row in ws.iter_rows():
            for cell in ws_row:
                if isinstance(cell.value, str) and "Caldat!" in cell.value:
                    cell.value = cell.value.replace("Caldat!", label0)

    # --- Overview sheet: one full-dataset chart per element, all sets overlaid ---
    ws_ov = wb.create_sheet("Overview", 0)
    ws_ov.cell(1, 1, f"XRF Calibration Overview \u2014 {run_date or ''}")
    ws_ov.cell(2, 1,
               f"{n_sets} calibration set(s): {', '.join(set_labels)}")

    CPR        = 3   # charts per row
    CHART_COLS = 14  # columns per chart slot (~12 cm)
    CHART_ROWS = 22  # rows per chart slot (~9 cm + gap)

    for chart_i, elem in enumerate(ELEMENT_CONFIG):
        ws_elem = wb[elem]
        ri, ci = divmod(chart_i, CPR)
        anchor = f"{_col_letter(1 + ci * CHART_COLS)}{3 + ri * CHART_ROWS}"

        chart = ScatterChart()
        chart.title = ELEMENT_FULL_NAMES[elem]
        chart.scatterStyle = "marker"
        chart.x_axis.title = "True Concentration (ppm)"
        chart.x_axis.numFmt = "0"
        chart.x_axis.tickLblPos = "nextTo"
        chart.x_axis.delete = False
        chart.y_axis.title = "Signal"
        chart.y_axis.numFmt = "0"
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.delete = False
        chart.width = 12
        chart.height = 9

        x_ref = Reference(ws_elem, min_col=7, min_row=2, max_row=8)

        y1_ref = Reference(ws_elem, min_col=8, min_row=2, max_row=8)
        chart.series.append(Series(y1_ref, xvalues=x_ref, title=set_labels[0]))

        for si in range(1, n_sets):
            col_sig = _signal_col(si)
            y_ref = Reference(ws_elem, min_col=col_sig, min_row=2, max_row=8)
            chart.series.append(Series(y_ref, xvalues=x_ref, title=set_labels[si]))

        ws_ov.add_chart(chart, anchor)

    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# 4. TABLE EXTRACTOR
# ---------------------------------------------------------------------------

def extract_benchling_tables(xlsx_path: pathlib.Path):
    """
    Read the Benchling Copyable Tables from the Samdat sheet.

    Returns:
        conc_rows:   [{'material','element','concentration','signal_quality','lod'}, ...]
        signal_rows: [{'material','element','signal_intensity'}, ...]
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Samdat"]
    rows = list(ws.iter_rows(values_only=True))

    conc_rows   = []
    signal_rows = []

    # Data starts at row index 3 (0-based), columns 4..8 and 11..13
    for row in rows[3:]:
        # Concentration table (cols E-I → indices 4-8)
        mat, elem, conc, quality, lod = row[4], row[5], row[6], row[7], row[8]
        if mat and elem:
            conc_rows.append({
                "material":      mat,
                "element":       elem,
                "concentration": None if conc == _DASH else conc,
                "signal_quality": quality,
                "lod":           lod,
            })

        # Signal intensity table (cols L-N → indices 11-13)
        mat2, elem2, sig = row[11], row[12], row[13]
        if mat2 and elem2:
            signal_rows.append({
                "material":         mat2,
                "element":          elem2,
                "signal_intensity": sig,
            })

    return conc_rows, signal_rows

# ---------------------------------------------------------------------------
# 5. BENCHLING UPLOADER
# ---------------------------------------------------------------------------

def _batches(items, size=100):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def _check_duplicate(benchling, xlsx_filename: str, entity_id: str = None):
    """Check for an existing summary result whose xlsx blob matches xlsx_filename.
    Filters by entity_id when provided to avoid scanning all historical results.
    Returns a dict {result_id, url} if found, otherwise None."""
    list_kwargs = {"schema_id": XRF_SUMMARY_SCHEMA}
    if entity_id:
        list_kwargs["entity_ids"] = [entity_id]

    pairs = []
    for page in benchling.assay_results.list(**list_kwargs):
        for r in page:
            field = (r.fields or {}).get(FIELD_SUM_REPORT)
            blob_id = getattr(field, "value", None)
            if blob_id:
                pairs.append((r.id, blob_id))

    if not pairs:
        return None

    # Bulk-fetch all matching blobs in one API call
    blobs = benchling.blobs.bulk_get([p[1] for p in pairs])
    blob_name_map = {b.id: b.name for b in blobs}

    for result_id, blob_id in pairs:
        if blob_name_map.get(blob_id) == xlsx_filename:
            return {
                "result_id": result_id,
                "url": f"{BENCHLING_URL}/assay-results/{result_id}",
            }
    return None


def find_entity_by_name(benchling, name: str):
    """Search Benchling custom entities by name prefix within the matching schema.
    Entity names in Benchling include lineage (e.g. 'Char-163-[Biomass-...]'),
    so we find the right schema from the name prefix then match startswith.
    Returns the first match or None.
    """
    # Determine schema from name prefix
    schema_id = None
    for prefix, sid in ENTITY_SCHEMA_MAP.items():
        if name.startswith(prefix):
            schema_id = sid
            break

    if schema_id:
        # Expand abbreviations so the prefix matches the full Benchling entity name
        search_prefix = name
        for abbrev, full in ENTITY_NAME_EXPANSIONS.items():
            if search_prefix.startswith(abbrev):
                search_prefix = full + search_prefix[len(abbrev):]
                break
        for page in benchling.custom_entities.list(schema_id=schema_id):
            for entity in page:
                if entity.name.startswith(search_prefix):
                    return entity
    else:
        # Fallback: search without schema filter, match by name prefix
        for page in benchling.custom_entities.list():
            for entity in page:
                if entity.name.startswith(name):
                    return entity
    return None


def upload_xrf_results(
    benchling,
    sample_name: str,
    conc_rows: list,
    signal_rows: list,
    xlsx_path: pathlib.Path,
    run_date: str,
    dry_run: bool = False,
    entity_id: str = None,
) -> dict:
    """Upload concentration, signal, and summary results to Benchling."""

    if dry_run:
        print(f"\n[DRY RUN] Sample: {sample_name}")
        print(f"  Concentration rows : {len(conc_rows)}")
        print(f"  Signal rows        : {len(signal_rows)}")
        print(f"  Report file        : {xlsx_path.name}")
        if entity_id:
            print(f"  Entity ID          : {entity_id}  (URL shown above during entity resolution)")
        else:
            print(f"  Entity link        : (none — will block live upload)")
        return {"dry_run": True}

    # Upload xlsx report as the summary blob
    blob = benchling.blobs.create_from_file(str(xlsx_path))

    # Concentration results — skip No Signal rows (no dropdown option exists)
    conc_creates = []
    for r in conc_rows:
        quality_id = SIGNAL_QUALITY_IDS.get(r["signal_quality"])
        if quality_id is None:
            continue
        f = {
            FIELD_CONC_SAMPLE:   {"value": entity_id},
            FIELD_CONC_ELEMENT:  {"value": CONC_ELEMENT_IDS[r["element"]]},
            FIELD_CONC_VALUE:    {"value": r["concentration"]},
            FIELD_CONC_QUALITY:  {"value": quality_id},
        }
        conc_creates.append(
            AssayResultCreate(
                schema_id=XRF_CONCENTRATION_SCHEMA,
                project_id=XRF_PROJECT_ID,
                fields=_fields(f),
            )
        )

    # Signal intensity results
    sig_creates = []
    for r in signal_rows:
        elem_id = SIG_ELEMENT_IDS.get(r["element"])
        if elem_id is None:
            continue  # element not in template/dropdown — skip
        f = {
            FIELD_SIG_SAMPLE:   {"value": [entity_id]},  # isMulti=True
            FIELD_SIG_ELEMENT:  {"value": elem_id},
            FIELD_SIG_VALUE:    {"value": r["signal_intensity"]},
        }
        sig_creates.append(
            AssayResultCreate(
                schema_id=XRF_SIGNAL_SCHEMA,
                project_id=XRF_PROJECT_ID,
                fields=_fields(f),
            )
        )

    # Summary result
    summary_fields = {
        FIELD_SUM_SAMPLE:  {"value": [entity_id]},  # isMulti=True
        FIELD_SUM_REPORT:  {"value": blob.id},
    }
    summary_create = AssayResultCreate(
        schema_id=XRF_SUMMARY_SCHEMA,
        project_id=XRF_PROJECT_ID,
        fields=_fields(summary_fields),
    )

    conc_ids = []
    for batch in _batches(conc_creates):
        task = benchling.assay_results.bulk_create(batch)
        conc_ids.extend(task.wait_for_response().assay_results)

    sig_ids = []
    for batch in _batches(sig_creates):
        task = benchling.assay_results.bulk_create(batch)
        sig_ids.extend(task.wait_for_response().assay_results)

    summary_id = (
        benchling.assay_results.bulk_create([summary_create])
        .wait_for_response()
        .assay_results[0]
    )

    return {
        "summary_result_id":     summary_id,
        "concentration_result_ids": conc_ids,
        "signal_result_ids":        sig_ids,
        "sample_name":              sample_name,
    }

# ---------------------------------------------------------------------------
# 6. MAIN ORCHESTRATOR
# ---------------------------------------------------------------------------

def upload_xrf_from_txt(
    txt_path: pathlib.Path,
    output_dir: pathlib.Path,
    benchling=None,
    dry_run: bool = False,
    entity_map: dict = None,
) -> list:
    """
    Full pipeline for one .txt file:
      1. Parse txt → samples + calibration signals
      2. Build calibration curves
      3. For each sample: process → generate xlsx → extract tables → upload
    Returns list of result dicts (one per sample).
    """
    txt_path   = pathlib.Path(txt_path)
    output_dir = pathlib.Path(output_dir)

    parsed = parse_xrf_txt(txt_path)
    samples       = parsed["samples"]
    cal_signals   = parsed["cal_standards"]
    cal_sets      = parsed["cal_sets"]
    all_elements  = parsed["all_elements"]
    run_date      = parsed["run_date"]

    if not cal_sets:
        raise ValueError(
            "No YP50F calibration standards found in txt file. "
            "Check that the Ident column contains YP50F-0 through YP50F-6."
        )

    print(f"\nParsed '{txt_path.name}'")
    print(f"  Calibration sets      : {len(cal_sets)}"
          f"  ({[sorted(cs.keys()) for cs in cal_sets]})")
    print(f"  Samples found         : {list(samples.keys())}")
    print(f"  Run date              : {run_date}")

    # Calibration-only file → generate comparison report and exit early
    if not samples:
        print("\n  No samples detected — generating calibration-only report.")
        xlsx_path = generate_xlsx_cal_report(cal_sets, all_elements, run_date, output_dir)
        print(f"  Calibration report written: {xlsx_path}")
        return [{"cal_only": True, "path": str(xlsx_path)}]

    calibration = build_calibration(cal_signals)

    all_results = []
    for sample_name, raw_signals in samples.items():
        print(f"\n  Processing: {sample_name}")
        processed = process_sample(raw_signals, calibration)

        # Print concentration preview
        for elem, r in processed.items():
            disp = r["display_conc"]
            print(f"    {elem:4s}: {str(disp):>20s}  [{r['quality']}]")

        xlsx_path = generate_xlsx_report(
            sample_name, cal_signals, raw_signals, all_elements, output_dir, run_date
        )
        print(f"  Report written: {xlsx_path}")

        entity_id = (entity_map or {}).get(sample_name)

        if benchling is not None:
            dupe = _check_duplicate(benchling, xlsx_path.name, entity_id=entity_id)
            if dupe:
                if dry_run:
                    print(f"  [DRY RUN] Existing upload found: {dupe['url']}")
                else:
                    print(f"  Duplicate detected — skipping. Existing result: {dupe['url']}")
                    all_results.append({"duplicate": True, "sample_name": sample_name, "url": dupe["url"]})
                    continue

        conc_rows, signal_rows = build_benchling_rows(sample_name, processed, raw_signals)
        # Filter signal rows to elements present in the template dropdown
        signal_rows = [r for r in signal_rows if r["element"] in SIG_ELEMENT_IDS]
        if not dry_run and entity_id is None:
            print(f"  Skipping Benchling upload for '{sample_name}' (no entity resolved).")
            all_results.append({"skipped": True, "sample_name": sample_name})
            continue

        result = upload_xrf_results(
            benchling,
            sample_name,
            conc_rows,
            signal_rows,
            xlsx_path,
            run_date,
            dry_run=dry_run,
            entity_id=entity_id,
        )
        all_results.append(result)

    return all_results
